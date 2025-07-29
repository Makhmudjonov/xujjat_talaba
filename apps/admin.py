from django.contrib import admin
from django.http import HttpResponse
import openpyxl
from simple_history.admin import SimpleHistoryAdmin
from django.contrib.auth.admin import UserAdmin
import urllib # Import UserAdmin directly

from apps.filter.dublikatfilter import DuplicateApplicationFilter
from apps.models import (
    Answer, Application, ApplicationFile, ApplicationItem, ApplicationType, ContractInfo, CustomAdminUser, Direction, Faculty,
    GPARecord, GroupHemis, Level, OdobAxloqStudent, Option, Question, Score, Section, SpecialApplicationStudent, Speciality, Student, Test, TestSession, University, # Make sure CustomAdminUser is imported
)

@admin.register(University)
class UniversityAdmin(SimpleHistoryAdmin):
    list_display = ('name',)

@admin.register(Student)
class StudentAdmin(SimpleHistoryAdmin):
    list_display = (
        'full_name',
        'student_id_number',
        'specialty',
        'group_hemis',
        'faculty',
        'group',
        'level',
        'toifa'
    )
    list_filter = ('faculty', 'level', 'gender', 'university', 'specialty', 'group_hemis','toifa')
    search_fields = ('full_name', 'student_id_number', 'email', 'phone')
    readonly_fields = ('student_id_number', 'full_name', 'email', 'phone')

 
@admin.register(Faculty)
class FacultyAdmin(SimpleHistoryAdmin):
    list_display = ('name', 'code')
    search_fields = ('name', 'code')

@admin.register(GPARecord)
class GPARecordAdmin(SimpleHistoryAdmin):
    list_display = ("student", "education_year", "level", "gpa", "credit_sum", "subjects", "debt_subjects")
    list_filter = ("education_year", "level", "can_transfer")
    search_fields = ("student__full_name",)

@admin.register(ContractInfo)
class ContractInfoAdmin(SimpleHistoryAdmin):
    list_display = ("student", "contract_number", "edu_year", "contract_sum", "debit", "credit")
    search_fields = ("student__full_name", "contract_number")

# Register your CustomAdminUser with its custom admin class
# Use @admin.register(CustomAdminUser) or admin.site.register(CustomAdminUser, CustomAdminUserAdmin)
# Do NOT use @admin.register(User)
# @admin.register(CustomAdminUser) # <--- Corrected this line
# class CustomAdminUserAdmin(UserAdmin): # <--- Inherit from UserAdmin directly
#     # These are default UserAdmin fields; you can customize them.
#     # If you have custom fields in CustomAdminUser, you'll need to add them here.
#     fieldsets = UserAdmin.fieldsets + (
#         ('Custom Fields', {'fields': ('sections', 'directions', 'faculties', 'levels', 'limit_by_course', 'allow_all_students')}),
#     )
#     add_fieldsets = UserAdmin.add_fieldsets + (
#         ('Custom Fields', {'fields': ('sections', 'directions', 'faculties', 'levels', 'limit_by_course', 'allow_all_students')}),
#     )
#     list_display = UserAdmin.list_display + ('limit_by_course', 'allow_all_students') # Add custom fields to list display
#     filter_horizontal = ('groups', 'user_permissions', 'sections', 'directions', 'faculties', 'levels') # Make many-to-many fields easier to manage


admin.site.register(Section)
admin.site.register(Direction)
# admin.site.register(Application, SimpleHistoryAdmin)


@admin.register(Score)
class ScoreAdmin(SimpleHistoryAdmin):
    list_display = ('id', 'get_student_full_name', 'get_direction', 'value', 'reviewer', 'scored_at')
    search_fields = ('item__application__student__full_name',)
    list_filter = (
        'scored_at',
        'item__direction',  # ✅ Direction bo‘yicha filter to‘g‘rilandi
    )

    def get_student_full_name(self, obj):
        return obj.item.application.student.full_name
    get_student_full_name.short_description = "Talaba"

    def get_direction(self, obj):
        return obj.item.direction.name if obj.item.direction else "-"
    get_direction.short_description = "Yo‘nalish"



    # def get_section(self, obj):
    #     direction = obj.item.application.direction
    #     return direction.section.name if direction and direction.section else "-"
    # get_section.short_description = "Bo‘lim (Section)"

class GroupLangFilter(admin.SimpleListFilter):
    title = 'Group Language'
    parameter_name = 'group_lang'

    def lookups(self, request, model_admin):
        return [
            ('O‘zbek', 'O‘zbek'),
            ('Rus', 'Rus'),
            ('Ingliz', 'Ingliz'),
        ]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(student__group_hemis__lang=self.value())
        return queryset

@admin.register(Application)
class ApplicationAdmin(SimpleHistoryAdmin):
    list_display = ('student', 'application_type', 'status', 'submitted_at', 'student__university','student__group_hemis', 'student__faculty', 'student__level')
    list_filter = ('application_type','student__university', 'student__university1', 'student__faculty','student__specialty',DuplicateApplicationFilter,GroupLangFilter)
    search_fields = ('student__full_name', 'student__student_id_number')  # misol uchun
    actions = ['export_as_excel', 'ijtimoiy_export_as_excel']
    
    def export_as_excel(self, request, queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Applications"

        # Sarlavha ustunlari — kerakli ma'lumotlar
        ws.append([
            "Student ID",
            "Full Name",
            "University",
            "Faculty",
            "Mutaxasislik",
            "Ta'lim shifri",
            "Hemis group",
            "Ta'lim tili",
            "Level",
            "Guruh",
            "Application Type",
            "Submitted At",
            "GPA",
            "GPA *16"
            "Score(s)",
        ])

        for app in queryset.select_related("student", "application_type").prefetch_related("items__score"):
            student = app.student
            items = app.items.all()

            # Har bir Application uchun ApplicationItemlar ketma-ket yoziladi
            # direction_names = ", ".join(str(item.direction.name) for item in items)
            comments = ", ".join(item.reviewer_comment or "-" for item in items)
            scores = ", ".join(str(item.score.value) if hasattr(item, "score") and item.score else "-" for item in items)

            ws.append([
                student.student_id_number,
                student.full_name,
                student.university1.name if student.university1 else "",
                student.faculty.name if student.faculty else "",
                student.specialty.name if student.specialty else "",
                student.specialty.code if student.specialty else "",
                student.group_hemis.name if student.group_hemis else "",
                student.group_hemis.lang if student.group_hemis else "",
                student.level.name if student.level else "",
                student.group if student.group else "",
                str(app.application_type),
                app.submitted_at.strftime('%Y-%m-%d %H:%M'),
                student.gpa or "",
                round(float(student.gpa) * 16, 3) or "",
                scores,
            ])

        selected_lang = request.GET.get("group_lang")
        lang_part = f"-{selected_lang}" if selected_lang else ""

        # Fayl nomini yasash
        filename = f"{student.university1.name}-{student.specialty.name}-{lang_part}-{student.specialty.code}-{app.application_type}.xlsx".replace("/", "-")
        filename_encoded = urllib.parse.quote(filename)
        
        # Excel response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename_encoded}"'
        wb.save(response)
        return response

    

    export_as_excel.short_description = "Excelga (barcha tafsilotlar bilan) eksport qilish"


    def ijtimoiy_export_as_excel(self, request, queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Applications"

        direction_names = set()
        for app in queryset.prefetch_related("items__direction"):
            for item in app.items.all():
                if item.direction:
                    direction_names.add(item.direction.name)
                direction_names.add("Jami ball")
                
        direction_names = sorted(direction_names)  # for consistent order

        # Sarlavha ustunlari — kerakli ma'lumotlar
        # 2. Header row
        headers = [
            "Student ID",
            "Full Name",
            "University",
            "Faculty",
            "Mutaxasislik",
            "Ta'lim shifri",
            "Hemis group",
            "Ta'lim tili",
            "Level",
            "Guruh",
            "Application Type",
            "Submitted At",
            "GPA",
            "GPA *16",
        ] + direction_names  # dynamically add direction columns

        ws.append(headers)

        for app in queryset.select_related("student", "application_type").prefetch_related("items__score"):
            student = app.student
            items = app.items.all()

            # Har bir Application uchun ApplicationItemlar ketma-ket yoziladi
            # direction_names = ", ".join(str(item.direction.name) for item in items)
            score_map = {
                item.direction.name: round(item.score.value * 0.2, 2) if item.direction.name.lower() == "Kitobxonlik madaniyati" and hasattr(item, "score") and item.score
                else (item.score.value if hasattr(item, "score") and item.score else "-")
                for item in items if item.direction
            }

            row = [
                student.student_id_number,
                student.full_name,
                student.university1.name if student.university1 else "",
                student.faculty.name if student.faculty else "",
                student.specialty.name if student.specialty else "",
                student.specialty.code if student.specialty else "",
                student.group_hemis.name if student.group_hemis else "",
                student.group_hemis.lang if student.group_hemis else "",
                student.level.name if student.level else "",
                student.group if student.group else "",
                str(app.application_type),
                app.submitted_at.strftime('%Y-%m-%d %H:%M') if app.submitted_at else "",
                student.gpa or "",
                round(float(student.gpa) * 16, 3) if student.gpa else "",
            ]

            # Append score values in the correct column order
            for dir_name in direction_names:
                row.append(score_map.get(dir_name, "-"))
            
            total_score = 0
            for dir_name in direction_names:
                value = score_map.get(dir_name, 0)
                if value == "-" or value == "" or value is None:
                    continue
                try:
                    value = float(value)
                    # if dir_name.lower() == "Kitobxonlik madaniyati":
                    #     value *= 0.2
                    total_score += value
                except ValueError:
                    pass

            row.append(round(total_score, 2))  # yoki butun son bo‘lsa: int(total_score)

            ws.append(row)

        selected_lang = request.GET.get("group_lang")
        lang_part = f"-{selected_lang}" if selected_lang else ""

        # Fayl nomini yasash
        filename = f"{student.university1.name}-{student.specialty.name}-{lang_part}-{student.specialty.code}-{app.application_type}.xlsx".replace("/", "-")
        filename_encoded = urllib.parse.quote(filename)
        
        # Excel response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename_encoded}"'
        wb.save(response)
        return response


    ijtimoiy_export_as_excel.short_description = "Ijtimoiy Excel (barcha tafsilotlar bilan) eksport qilish"

@admin.register(ApplicationType)
class ApplicationTypeAdmin(SimpleHistoryAdmin):
    list_display = ('name', 'min_gpa')
    # list_filter = ('access_type',)


@admin.register(Speciality)
class SpecialityAdmin(admin.ModelAdmin):
    list_display = ('name', 'code', 'hemis_id', 'university')
    search_fields = ('name', 'code', 'hemis_id')
    list_filter = ('university',)

    def __str__(self):
        return f"{self.name}"

@admin.register(GroupHemis)
class GroupHemisAdmin(admin.ModelAdmin):
    list_display = ('name', 'hemis_id', 'lang')
    search_fields = ('name', 'hemis_id', 'lang')
    list_filter = ('lang',)

    def __str__(self):
        return f"{self.name}"


@admin.register(SpecialApplicationStudent)
class SpecialApplicationStudentAdmin(SimpleHistoryAdmin):
    list_display = ('hemis_id', 'application_type', 'student')
    search_fields = ('hemis_id',)
    list_filter = ('application_type',)

# @admin.register(Student)
# class StudentAdmin(SimpleHistoryAdmin):
#     list_display = ('hemis_id', 'last_name', 'first_name', 'gpa')
#     search_fields = ('hemis_id', 'last_name')


from django.utils.html import escape
from django.conf import settings
from urllib.parse import urljoin
import openpyxl
from django.http import HttpResponse


import zipfile
import io
import os
from django.utils.text import slugify
from django.conf import settings

@admin.register(ApplicationItem)
class ApplicationItemAdmin(SimpleHistoryAdmin):
    list_display = ("id", "get_student_name", "get_level", "file", "direction")
    search_fields = ('application__student__full_name',)
    list_filter = ('direction', 'application__student__level', 'application__student__faculty')
    actions = ["export_as_excel", "download_all_files_as_zip"]

    def get_student_name(self, obj):
        return obj.application.student.full_name
    get_student_name.short_description = "Talaba"

    def get_level(self, obj):
        return obj.application.student.level.name
    get_level.short_description = "Bosqich (Level)"

    def export_as_excel(self, request, queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Application Items"

        # Excel ustun sarlavhalari
        ws.append([
            "ID",
            "Talaba",
            "Bosqich",
            "Fakultet",
            "Yo'nalish",
            "Sarlavha",
            "Main Fayl (nomi)",
            "Main Fayl (havola)",
            "Talaba izohi",
            "Arizani topshirish vaqti",
            "O'quv yili",
            "Qo‘shimcha Fayl (havola)",
            "Qo‘shimcha Fayl izohi",
            "Yuklangan vaqti",
        ])

        for item in queryset.select_related("application__student", "direction", "application__student__level", "application__student__faculty").prefetch_related("files"):
            student = item.application.student

            main_file_name = item.file.name if item.file else ''
            main_file_url = urljoin(
                request.build_absolute_uri('/'),
                f"{settings.MEDIA_URL}{main_file_name}"
            ) if main_file_name else ''

            # Agar hech qanday ApplicationFile bo‘lmasa, hech bo‘lmasa bitta qator chiqsin
            related_files = list(item.files.all())
            if not related_files:
                ws.append([
                    item.id,
                    student.full_name,
                    student.level.name if student.level else '',
                    student.faculty.name if student.faculty else '',
                    item.direction.name,
                    item.title,
                    main_file_name,
                    main_file_url,
                    item.student_comment or '',
                    '', '', '', ''
                ])
            else:
                for f in related_files:
                    file_url = urljoin(
                        request.build_absolute_uri('/'),
                        f"{settings.MEDIA_URL}{f.file.name}"
                    ) if f.file else ''
                    ws.append([
                        item.id,
                        student.full_name,
                        student.level.name if student.level else '',
                        student.faculty.name if student.faculty else '',
                        item.direction.name,
                        item.title,
                        main_file_name,
                        main_file_url,
                        item.student_comment or '',
                        item.application.submitted_at.strftime('%Y-%m-%d %H:%M') if item.application.submitted_at else '',
                        f.section.name if f.section else '',
                        file_url,
                        f.comment or '',
                        f.uploaded_at.strftime('%Y-%m-%d %H:%M'),
                    ])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=application_items.xlsx"
        wb.save(response)
        return response

    export_as_excel.short_description = "Excelga yuklab olish (fayllar bilan)"

    def download_all_files_as_zip(self, request, queryset):
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            for item in queryset.prefetch_related("files").select_related("application__student"):
                student = item.application.student
                student_name = slugify(student.full_name)
                item_folder = f"{student_name}_{item.id}/"

                # Asosiy fayl
                if item.file:
                    main_file_path = item.file.path
                    if os.path.exists(main_file_path):
                        zip_file.write(
                            main_file_path,
                            arcname=os.path.join(item_folder, os.path.basename(main_file_path))
                        )

                # Qo‘shimcha fayllar (ApplicationFile)
                for f in item.files.all():
                    if f.file:
                        extra_file_path = f.file.path
                        if os.path.exists(extra_file_path):
                            section_name = slugify(f.section.name) if f.section else "unknown_section"
                            filename = f"{section_name}_{os.path.basename(f.file.name)}"
                            zip_file.write(
                                extra_file_path,
                                arcname=os.path.join(item_folder, "extra_files", filename)
                            )

        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer, content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename=application_items_files.zip'
        return response

    download_all_files_as_zip.short_description = "Tanlangan fayllarni ZIP qilib yuklab olish"



    


@admin.register(CustomAdminUser)
class CustomAdminUserAdmin(UserAdmin):
    model = CustomAdminUser
    list_display = ("username", "email", "role", "is_active", "is_staff", "can_score")
    list_filter = ("role", "is_active", "is_staff", "can_score")
    search_fields = ("username", "email", "first_name", "last_name")
    ordering = ("id",)

    fieldsets = (
        (None, {"fields": ("username", "password")}),
        ("Shaxsiy ma'lumotlar", {"fields": ("first_name", "last_name", "email")}),
        ("Ruxsatlar", {
            "fields": (
                "is_active", "is_staff", "is_superuser",
                "groups", "user_permissions",
            )
        }),
        ("Admin rollari", {
            "fields": (
                "role", 'full_name', "sections", "faculties", "directions", "levels",
                "limit_by_course", "allow_all_students", "can_score", 'university1', 
            )
        }),
        ("Muhim sanalar", {"fields": ("last_login", "date_joined")}),
    )

    add_fieldsets = (
        (None, {
            "classes": ("wide",),
            "fields": (
                "username", "password1", "password2", "email",
                "role", "is_staff", "is_active",
                "sections", "faculties", "directions", "levels",
                "limit_by_course", "allow_all_students", "can_score", 'university1' # ✅ Qo‘shildi
            ),
        }),
    )



@admin.register(ApplicationFile)
class ApplicationFileAdmin(SimpleHistoryAdmin):
    list_display = ('id', 'get_application', 'comment')

    def get_application(self, obj):
        return obj.item.application
    get_application.short_description = "Application"




#test admin
class AnswerInline(admin.TabularInline):
    model = Answer
    extra = 2


@admin.register(Answer)
class AnswerAdmin(SimpleHistoryAdmin):
    list_display = ('id', 'question', 'is_correct')
    list_filter = ('is_correct', 'session')

class OptionInline(admin.TabularInline):
    model = Option
    extra = 4  # savolga default 4 javob varianti ko‘rsatiladi

class QuestionAdmin(SimpleHistoryAdmin):
    inlines = [OptionInline]
    list_display = ('text',)

admin.site.register(Question, QuestionAdmin)

@admin.register(Option)
class OptionAdmin(SimpleHistoryAdmin):
    list_display = ('question', 'label', 'text', 'is_correct')
    list_filter = ('is_correct',)


@admin.register(Test)
class TestAdmin(SimpleHistoryAdmin):
    list_display = ("title", "question_count", "time_limit", "start_time", "created_at")  # 🆕 start_time qo‘shildi
    search_fields = ("title",)
    list_filter = ("created_at", "levels", "start_time")  # 🆕
    filter_horizontal = ("levels",)

    fieldsets = (
        (None, {
            "fields": ("title", "question_count", "time_limit", "start_time", "levels")  # 🆕
        }),
    )


@admin.register(TestSession)
class TestSessionAdmin(SimpleHistoryAdmin):
    list_display = ("id", "student", "test", "started_at", "finished_at", "score")
    list_filter = ("test", "student")
    search_fields = ("student__full_name", "student__student_id_number", "test__title")
    readonly_fields = ("started_at", "finished_at", "score", "correct_answers", "total_questions")

    def get_queryset(self, request):
        return super().get_queryset(request).select_related("student", "test")


@admin.register(OdobAxloqStudent)
class OdobAxloqStudentAdmin(SimpleHistoryAdmin):
    list_display = ('id', 'hemis_id', 'sabab')
    search_fields = ('hemis_id',)
    list_filter = ('sabab',)



@admin.register(Level)
class LevelAdmin(SimpleHistoryAdmin):
    list_display = ('id', 'code', 'name')
    search_fields = ('code', 'name')