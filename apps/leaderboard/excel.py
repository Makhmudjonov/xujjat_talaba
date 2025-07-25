from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
from django.db.models import Sum, FloatField, Q, F, Value as V
from django.db.models.functions import Coalesce

from apps.models import Student, TestSession

class LeaderboardExportAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        admin = request.user
        students = Student.objects.prefetch_related(
            'applications__items__score',
            'applications__items__direction',
            'gpa_records',
            'faculty',
            'level'
        ).all()

        # Admin filtrlar
        if hasattr(admin, "role") and admin.role != "student":
            if admin.university1:
                students = students.filter(university1=admin.university1)
            if admin.faculties.exists():
                students = students.filter(faculty__in=admin.faculties.all())
            if admin.levels.exists():
                students = students.filter(level__in=admin.levels.all())
            if admin.directions.exists():
                students = students.filter(applications__items__direction__in=admin.directions.all())

        # Query params
        faculty_id = request.GET.get("faculty")
        level_id = request.GET.get("level")
        course = request.GET.get("course")
        university = request.GET.get("university")
        toifa_param = request.GET.get("toifa")

        if faculty_id:
            students = students.filter(faculty_id=faculty_id)
        if level_id:
            students = students.filter(level_id=level_id)
        if course:
            students = students.filter(group__icontains=course)
        if university:
            students = students.filter(university1_id=university)
        if toifa_param is not None:
            if toifa_param.lower() in ["true", "1"]:
                students = students.filter(toifa=True)
            elif toifa_param.lower() in ["false", "0"]:
                students = students.filter(toifa=False)

        # Annotatsiya: GPA + score
        students = students.annotate(
            gpa_sum=Coalesce(Sum('gpa_records__gpa'), V(0.0), output_field=FloatField()),
            gpa_count=Coalesce(Sum(V(1), filter=~Q(gpa_records=None)), V(0.0001), output_field=FloatField()),
            score_sum=Coalesce(Sum('applications__items__score__value'), V(0.0), output_field=FloatField()),
        ).annotate(
            total_score=F('gpa_sum') / F('gpa_count') + F('score_sum')
        ).order_by('-total_score')

        return self.export_to_excel(students)

    def export_to_excel(self, students):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Leaderboard"

        headers = [
            "ID", "Full Name", "Universitet", "Fakultet", "Level", "Group", "Total Score", "Toifa"
        ]
        ws.append(headers)

        for student in students:
            total_score = 0.0

            # GPA bo'yicha eng oxirgi yozuv
            latest_gpa = student.gpa_records.order_by('-created_at').first()
            if latest_gpa:
                gpa = float(latest_gpa.gpa)
                gpa_score_map = {
                    5.0: 10.0, 4.9: 9.7, 4.8: 9.3, 4.7: 9.0,
                    4.6: 8.7, 4.5: 8.3, 4.4: 8.0, 4.3: 7.7,
                    4.2: 7.3, 4.1: 7.0, 4.0: 6.7, 3.9: 6.3,
                    3.8: 6.0, 3.7: 5.7, 3.6: 5.3, 3.5: 5.0,
                }
                total_score += gpa_score_map.get(round(gpa, 1), 0.0)

            for app in student.applications.all():
                for item in app.items.all():
                    direction_name = item.direction.name if item.direction else ""

                    if direction_name == 'Kitobxonlik madaniyati':
                        # Kitobxonlik uchun test natijasidan olish
                        test = getattr(item.direction, 'test', None)
                        if test:
                            session = TestSession.objects.filter(student=student, test=test).order_by('-finished_at').first()
                            if session and session.correct_answers is not None:
                                total_score += round(session.correct_answers * 20 / 25, 2)

                    elif direction_name == 'Talabaning akademik o‘zlashtirishi':
                        pass  # oldin GPA dan qo‘shilgan

                    else:
                        if hasattr(item, 'score') and item.score:
                            total_score += float(item.score.value)

            ws.append([
                student.id,
                student.full_name,
                student.university1.name if student.university1 else "",
                student.faculty.name if student.faculty else "",
                student.level.name if student.level else "",
                student.group,
                round(total_score, 2),
                "Nogiron" if getattr(student, 'toifa', False) else "Oddiy"
            ])

        # Auto column width
        for col_num, _ in enumerate(ws.columns, 1):
            ws.column_dimensions[get_column_letter(col_num)].auto_size = True

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=leaderboard.xlsx'
        wb.save(response)
        return response

