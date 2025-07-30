import openpyxl
from openpyxl.utils import get_column_letter
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.http import HttpResponse
from io import BytesIO
from drf_yasg import openapi
from drf_yasg.utils import swagger_auto_schema

from apps.gpaStudent.studentList import IsAdminUser
from apps.models import Application, TestSession

class ApplicationExportExcelAPIView(APIView):
    permission_classes = [IsAuthenticated, IsAdminUser]

    def get_queryset(self):
        user = self.request.user

        qs = Application.objects.prefetch_related(
            "items__direction__section",
            "items__files",
            "items__score",
            "student__faculty",
            "student__level",
        ).select_related("application_type", "student")

        if user.university1:
            qs = qs.filter(student__university1=user.university1)

        if user.faculties.exists():
            qs = qs.filter(student__faculty__in=user.faculties.all())

        if user.levels.exists():
            qs = qs.filter(student__level__in=user.levels.all())

        if user.directions.exists():
            qs = qs.filter(items__direction__in=user.directions.all())

        status = self.request.query_params.get("status")
        if status:
            qs = qs.filter(status=status)

        student = self.request.query_params.get("student")
        if student:
            qs = qs.filter(student__full_name__icontains=student)

        return qs.distinct()

    @swagger_auto_schema(
        operation_summary="Application ro'yxatini filterga qarab Excel formatda yuklab olish",
        manual_parameters=[
            openapi.Parameter("status", openapi.IN_QUERY, description="Status (pending, accepted, rejected)", type=openapi.TYPE_STRING),
            openapi.Parameter("student", openapi.IN_QUERY, description="Talaba ismi bo‘yicha qidiruv", type=openapi.TYPE_STRING),
        ],
    )
    def get(self, request, *args, **kwargs):
        queryset = self.get_queryset()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Applications"

        # Dinamik direction nomlarini yig'ish
        direction_names = set()
        for app in queryset.prefetch_related("items__direction"):
            for item in app.items.all():
                if item.direction:
                    direction_names.add(item.direction.name)
        direction_names = sorted(direction_names)

        # Sarlavha
        headers = [
            "ID", "Student", "Faculty", "Level", "Application Type", "Status"
        ] + direction_names + ["Jami ball", "Jami ball * 0.2"]

        ws.append(headers)

        def get_gpa_score(gpa):
            if gpa is None:
                return 0.0
            gpa_score_map = {
                5.0: 10.0, 4.9: 9.7, 4.8: 9.3, 4.7: 9.0,
                4.6: 8.7, 4.5: 8.3, 4.4: 8.0, 4.3: 7.7,
                4.2: 7.3, 4.1: 7.0, 4.0: 6.7, 3.9: 6.3,
                3.8: 6.0, 3.7: 5.7, 3.6: 5.3, 3.5: 5.0,
            }
            return gpa_score_map.get(round(gpa, 2), 0.0)

        for app in queryset.select_related("student", "application_type").prefetch_related("items__score", "student__gpa_records"):
            student = app.student
            score_map = {}
            total_score = 0

            for item in app.items.all():
                if not item.direction:
                    continue
                dir_name = item.direction.name.lower()

                if dir_name == "kitobxonlik madaniyati":
                    try:
                        test = TestSession.objects.filter(student=student).first()
                        val = round(float(test.score) * 0.2, 2)
                        score_map[item.direction.name] = val
                        total_score += val
                    except:
                        score_map[item.direction.name] = "Mavjud emas"
                elif dir_name == "talabaning akademik o‘zlashtirishi":
                    gpa = student.gpa or None
                    gpa_score = get_gpa_score(round(float(gpa), 1)) if gpa else 0
                    score_map[item.direction.name] = gpa_score if gpa_score else "-"
                    total_score += gpa_score
                else:
                    val = item.score.value if hasattr(item, "score") and item.score else "-"
                    score_map[item.direction.name] = val
                    if isinstance(val, (float, int)):
                        total_score += val

            # Asosiy ustunlar
            row = [
                app.id,
                student.full_name,
                student.faculty.name if student.faculty else "",
                student.level.name if student.level else "",
                app.application_type.name if app.application_type else "Unknown",
                app.status if app.status else "Unknown",
            ]

            # Har bir direction uchun qiymat
            for dir_name in direction_names:
                row.append(score_map.get(dir_name, "-"))

            row.append(round(total_score, 2))
            row.append(round(total_score * 0.2, 2))

            ws.append(row)

        # Auto width
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 3

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = "applications.xlsx"
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{urllib.parse.quote(filename)}"'
        return response
