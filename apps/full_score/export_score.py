import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from rest_framework.views import APIView
from apps.gpaStudent.studentList import IsAdminUser
from apps.models import Student
from apps.serializers import StudentCombinedScoreSerializer


class ExportStudentScoreExcelView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        user = request.user

        # Boshlang'ich queryset
        queryset = Student.objects.prefetch_related(
            'applications__items__score',
            'applications__items__direction',
            'gpa_records',
            'faculty',
            'level'
        ).all()

        # Faqat admin, dekan, kichik_admin rollari uchun filtering
        if user.role in ['admin', 'dekan', 'kichik_admin']:
            if user.university1:
                queryset = queryset.filter(university=user.university1)
            if user.faculties.exists():
                queryset = queryset.filter(faculty__in=user.faculties.all())
            if user.levels.exists():
                queryset = queryset.filter(level__in=user.levels.all())

        # Superuser barcha ma’lumotni ko‘ra oladi (filterlashsiz)
        # (Agar `is_superuser` tekshiruvi kerak bo‘lsa, u holda yuqoridagi filterlarni o'tkazib yuborishingiz mumkin)

        serializer = StudentCombinedScoreSerializer(queryset, many=True, context={'request': request})

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Student Scores"

        headers = ['ID', 'FISH', 'Fakultet', 'Kurs', 'Guruh', 'Toifa', 'GPA Ball/*16', 'Ijtimoiy Index natijasi', 'Jami ball(GPA ball + (Ijtimoiy Index natijasi * 0.2)']
        ws.append(headers)

        for row in serializer.data:
            total_score_data = row.get('total_score', {})
            ws.append([
                row.get('id'),
                row.get('full_name'),
                row.get('faculty'),
                row.get('course'),
                row.get('group'),
                "Bor" if row.get('toifa') else "Yo‘q",
                total_score_data.get('gpa', 0),
                total_score_data.get('score_total', 0),
                total_score_data.get('total', 0),
            ])

        # Auto column width
        for col in ws.columns:
            max_length = 0
            column = col[0].column  # 1-based index
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[get_column_letter(column)].width = max_length + 4

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=student_scores.xlsx'
        wb.save(response)
        return response
